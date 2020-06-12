#!/usr/bin/env python
#
# Copyright (c) $today.year Moises Martinez (Sngular). All Rights Reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
# ==============================================================================

from data import Tree
from .step import Step
from util import Tools
from datetime import datetime
from .inventory import Inventory
from scenario import Operation
from openpyxl import Workbook
from openpyxl import drawing
from openpyxl.styles import Alignment
from models import HarvestModel
from constants import OUTPUT_FILE_BASE
from constants import OUTPUT_EXTENSION
from constants import OUTPUT_NAMES
from constants import PLOT_VARIABLE_NAMES

import logging
import os
import i18n
import json

from dask import delayed

XLSX = 0
JSON = 1


class Simulation:

    def __init__(self, date=datetime.now()):

        self.__date = date
        self.__steps = list()

    def add_step(self, step_id, inventory: Inventory, operation: Operation, model):

        min_age = 0
        max_age = 0

        if len(self.__steps) == 0:
            age = operation.get_variable('init')
        else:
            last = self.get_last()
            age = last.age + operation.get_variable('time')
            min_age = operation.get_variable('min_age')
            max_age = operation.get_variable('max_age')

        model_name = ''

        if isinstance(model, HarvestModel):
            model_name = i18n.t('simanfor.general.' + model.name)

        self.__steps.append(Step(step_id, inventory, operation.type, operation.description, 
                                 age, min_age, max_age, operation, model_name))

    def get_step(self, position):
        if position < len(self.__steps):
            return self.__steps[position]
        return None

    def get_first_step(self):
        return self.get_step(0)

    def get_last(self):
        if len(self.__steps) == 0:
            return 0
        return self.__steps[len(self.__steps)-1]


    def generate_json_file(self, name: str, file_path, plot_id: str):

        # plots = self.__steps[0].get_plot_ids()

        general = dict()

        summary = dict()
        steps = dict()

        summary[i18n.t('simanfor.general.main_specie')] = ''
        summary[i18n.t('simanfor.general.main_specie')] = ''
        summary[i18n.t('simanfor.general.forest')] = ''
        summary[i18n.t('simanfor.general.study_area')] = ''
        summary[i18n.t('simanfor.general.model')] = ''
        summary[i18n.t('simanfor.general.scenario')] = name
        summary[i18n.t('simanfor.general.inventory')] = ''
        summary[i18n.t('simanfor.general.template')] = ''

        row = 1

        for step in self.__steps:
            steps[row] = step.to_json(plot_id, OUTPUT_NAMES, row)
            row += 1

        general[i18n.t('simanfor.general.' + DEFAULT_EXCEL_OUTPUT_FILE_STRUCTURE[0])] = summary
        general[i18n.t('simanfor.general.' + DEFAULT_EXCEL_OUTPUT_FILE_STRUCTURE[1])] = steps

        with open(file_path, 'w') as outfile:
            json.dump(general, outfile)


    def resumen_merged_header(self, sheet, cell_range, row, col, val):
        
        sheet.merge_cells(cell_range)
        cell = sheet.cell(row = row, column = col)
        cell.value = val
        cell.alignment = Alignment(horizontal = 'center')
        cell.font = cell.font.copy(bold = True)        
    

    def generate_xslt_file(self, name: str, labels: dict, file_path: str, plot, modelo: str, zip_compression: bool = False, 
                           type: int = JSON, decimals: int = 2):

        Tools.print_log_line('Generating xslf file for plot ' + str(plot.id), logging.INFO)

        workbook = Workbook()

        worksheet = workbook.active
        worksheet.title = labels['simanfor.general.Summary']

        try:
            logo_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "..", "files", "logo.png")
            logo = drawing.image.Image(logo_path)
    
            logo.anchor = 'A1'
            logo.width = 190
            logo.height = 60
            worksheet.add_image(logo)
        except:
            Tools.print_log_line("Couldn't find logo image.", logging.INFO)

        worksheet.column_dimensions["D"].width = 20
        worksheet.cell(row=1, column=4).value = labels['simanfor.general.study_area']
        worksheet.cell(row=2, column=4).value = labels['simanfor.general.forest']
        worksheet.cell(row=3, column=4).value = labels['simanfor.general.main_specie']
        worksheet.cell(row=4, column=4).value = labels['simanfor.general.specie_ifn_id']
        
        worksheet.cell(row=1, column=5).value = plot.study_area
        worksheet.cell(row=2, column=5).value = plot.forest
        worksheet.cell(row=3, column=5).value = plot.main_specie
        worksheet.cell(row=4, column=5).value = plot.specie_ifn_id        
        
        worksheet.column_dimensions["J"].width = 16
        worksheet.column_dimensions["K"].width = 20
        
        worksheet.cell(row=1, column=10).value = labels['simanfor.general.inventory']
        worksheet.cell(row=2, column=10).value = labels['simanfor.general.plot']
        worksheet.cell(row=3, column=10).value = labels['simanfor.general.model']
        worksheet.cell(row=4, column=10).value = labels['simanfor.general.scenario']

        worksheet.cell(row=1, column=11).value = str(int(plot.inventory_id))
        worksheet.cell(row=2, column=11).value = str(plot.id)
        worksheet.cell(row=3, column=11).value = modelo
        worksheet.cell(row=4, column=11).value = name        

        self.resumen_merged_header(worksheet, 'C6:F6', 6, 3, labels['simanfor.general.merged_header_1'])
        self.resumen_merged_header(worksheet, 'G6:I6', 6, 7, labels['simanfor.general.merged_header_2'])
        self.resumen_merged_header(worksheet, 'J6:M6', 6, 10, labels['simanfor.general.merged_header_3'])
        self.resumen_merged_header(worksheet, 'N6:P6', 6, 14, labels['simanfor.general.merged_header_4'])

        cell = worksheet.cell(row=6, column=17)
        cell.value = labels['simanfor.general.merged_header_5']
        cell.font = cell.font.copy(bold = True)        
        
        worksheet.cell(row=7, column=1).value = labels['simanfor.plot.AGE']
        worksheet.cell(row=7, column=2).value = labels['simanfor.general.measure_1']
        worksheet.cell(row=7, column=3).value = labels['simanfor.general.measure_2']
        worksheet.cell(row=7, column=4).value = labels['simanfor.general.measure_3']
        worksheet.cell(row=7, column=5).value = labels['simanfor.general.measure_4']
        worksheet.cell(row=7, column=6).value = labels['simanfor.general.measure_5']
        worksheet.cell(row=7, column=7).value = labels['simanfor.general.measure_6']
        worksheet.cell(row=7, column=8).value = labels['simanfor.general.measure_7']
        worksheet.cell(row=7, column=9).value = labels['simanfor.general.measure_8']
        worksheet.cell(row=7, column=10).value = labels['simanfor.general.measure_9']
        worksheet.cell(row=7, column=11).value = labels['simanfor.general.measure_10']
        worksheet.cell(row=7, column=12).value = labels['simanfor.general.measure_11']
        worksheet.cell(row=7, column=13).value = labels['simanfor.general.measure_12']
        worksheet.cell(row=7, column=14).value = labels['simanfor.general.measure_13']
        worksheet.cell(row=7, column=15).value = labels['simanfor.general.measure_14']
        worksheet.cell(row=7, column=16).value = labels['simanfor.general.measure_15']
        worksheet.cell(row=7, column=17).value = labels['simanfor.general.measure_16']
        worksheet.cell(row=7, column=18).value = i18n.t('simanfor.general.measure_17')

        ws_parcelas = workbook.create_sheet(labels['simanfor.general.Plots'])
        
        for j in range(len(OUTPUT_NAMES)):
            ws_parcelas.cell(row = 1, column = j + 1).value = labels['simanfor.plot.' + OUTPUT_NAMES[j]]        
        for j in range(len(PLOT_VARIABLE_NAMES)):
            ws_parcelas.cell(row = 1, column = j + 1 + len(OUTPUT_NAMES)).value = labels['simanfor.plot.' + PLOT_VARIABLE_NAMES[j]]        

        step_count = 1
        row = 8
        summary_row = 8

        for step in self.__steps:

            next_step = None if step_count >= len(self.__steps) else self.__steps[step_count]

            next_operation = None if next_step == None else next_step._Step__full_operation          
            summary_row = step.to_xslt(labels, workbook, plot.id, row, next_step, next_operation, 
                                       summary_row, decimals)
            row += 1
            step_count += 1

        workbook.close()
        workbook.save(file_path)
        
        return plot


    def get_labels(self):
 
        labels = dict()

        # 
        # general
        # 
        labels['simanfor.general.Summary'] = i18n.t('simanfor.general.Summary')      
        
        labels['simanfor.general.main_specie'] = i18n.t('simanfor.general.main_specie')
        labels['simanfor.general.forest'] = i18n.t('simanfor.general.forest')
        labels['simanfor.general.study_area'] = i18n.t('simanfor.general.study_area')
        labels['simanfor.general.model'] = i18n.t('simanfor.general.model')

        labels['simanfor.general.specie_ifn_id'] = i18n.t('simanfor.general.specie_ifn_id')
        labels['simanfor.general.scenario'] = i18n.t('simanfor.general.scenario')
        labels['simanfor.general.inventory'] = i18n.t('simanfor.general.inventory')
        labels['simanfor.general.plot'] = i18n.t('simanfor.general.plot')

        labels['simanfor.general.merged_header_1'] = i18n.t('simanfor.general.merged_header_1')
        labels['simanfor.general.merged_header_2'] = i18n.t('simanfor.general.merged_header_2')
        labels['simanfor.general.merged_header_3'] = i18n.t('simanfor.general.merged_header_3')
        labels['simanfor.general.merged_header_4'] = i18n.t('simanfor.general.merged_header_4')
        labels['simanfor.general.merged_header_5'] = i18n.t('simanfor.general.merged_header_5')

        labels['simanfor.plot.AGE'] = i18n.t('simanfor.plot.AGE')
        labels['simanfor.general.measure_1'] = i18n.t('simanfor.general.measure_1')
        labels['simanfor.general.measure_2'] = i18n.t('simanfor.general.measure_2')
        labels['simanfor.general.measure_3'] = i18n.t('simanfor.general.measure_3')
        labels['simanfor.general.measure_4'] = i18n.t('simanfor.general.measure_4')
        labels['simanfor.general.measure_5'] = i18n.t('simanfor.general.measure_5')
        labels['simanfor.general.measure_6'] = i18n.t('simanfor.general.measure_6')
        labels['simanfor.general.measure_7'] = i18n.t('simanfor.general.measure_7')
        labels['simanfor.general.measure_8'] = i18n.t('simanfor.general.measure_8')
        labels['simanfor.general.measure_9'] = i18n.t('simanfor.general.measure_9')
        labels['simanfor.general.measure_10'] = i18n.t('simanfor.general.measure_10')
        labels['simanfor.general.measure_11'] = i18n.t('simanfor.general.measure_11')
        labels['simanfor.general.measure_12'] = i18n.t('simanfor.general.measure_12')
        labels['simanfor.general.measure_13'] = i18n.t('simanfor.general.measure_13')
        labels['simanfor.general.measure_14'] = i18n.t('simanfor.general.measure_14')
        labels['simanfor.general.measure_15'] = i18n.t('simanfor.general.measure_15')
        labels['simanfor.general.measure_16'] = i18n.t('simanfor.general.measure_16')
        labels['simanfor.general.measure_17'] = i18n.t('simanfor.general.measure_17')

        labels['simanfor.general.Plots'] = i18n.t('simanfor.general.Plots')
        labels['simanfor.general.Node'] = i18n.t('simanfor.general.Node')
        labels['simanfor.general.Trees'] = i18n.t('simanfor.general.Trees')
        # labels['simanfor.general.template'] = i18n.t('simanfor.general.template')

        labels['simanfor.general.LOAD'] = i18n.t('simanfor.general.LOAD')
        labels['simanfor.general.INIT'] = i18n.t('simanfor.general.INIT')
        labels['simanfor.general.EXECUTION'] = i18n.t('simanfor.general.EXECUTION')
        labels['simanfor.general.HARVEST'] = i18n.t('simanfor.general.HARVEST')
        labels['simanfor.general.Cut Down by Tallest'] = i18n.t('simanfor.general.Cut Down by Tallest')
        labels['simanfor.general.Cut Down by Smallest'] = i18n.t('simanfor.general.Cut Down by Smallest')
        labels['simanfor.general.Systematics cut down'] = i18n.t('simanfor.general.Systematics cut down')
        labels['simanfor.general.Empty'] = i18n.t('simanfor.general.Empty')
        labels['simanfor.general.Percent of trees'] = i18n.t('simanfor.general.Percent of trees')
        labels['simanfor.general.Volumen'] = i18n.t('simanfor.general.Volumen')
        labels['simanfor.general.Area'] = i18n.t('simanfor.general.Area')

        # 
        # plot
        # 
        for j in range(len(OUTPUT_NAMES)):
            labels['simanfor.plot.' + OUTPUT_NAMES[j]] = i18n.t('simanfor.plot.' + OUTPUT_NAMES[j])
        for j in range(len(PLOT_VARIABLE_NAMES)):
            labels['simanfor.plot.' + PLOT_VARIABLE_NAMES[j]] = i18n.t('simanfor.plot.' + PLOT_VARIABLE_NAMES[j])

        # 
        # tree
        # 
        for i in range(len(Tree.variables_names())):
            labels['simanfor.tree.' + Tree.variables_names()[i]] = i18n.t('simanfor.tree.' + Tree.variables_names()[i])
        
        return labels


    def generate_results(self, name: str, file_path: str, modelo: str, type: int = XLSX, zip_compression: bool = False, 
                         decimals: int = 2):

        labels = self.get_labels()

        plots = self.get_first_step().inventory.plots

        for plot in plots:

            if isinstance(plot.id, str): # json input
                filename = file_path + OUTPUT_FILE_BASE + os.path.split(plot.id)[1] + '.' + OUTPUT_EXTENSION[type]
            else:
                filename = file_path + OUTPUT_FILE_BASE + str(plot.id) + '.' + OUTPUT_EXTENSION[type]

            if type == XLSX:

                self.generate_xslt_file(
                    name, labels,
                    filename,
                    plot, modelo, decimals=decimals)

            # else:
            #     self.generate_json_file(
            #         name,
            #         current_folder + OUTPUT_FILE_BASE + str(plot.id) + '.' + OUTPUT_EXTENSION[type],
            #         plot.id)


    def print_plots(self, plots: list):
        
        # count = 0
        # for plot in plots:
        #     count += 1
        #     # self.print_plot(plot)
            
        return len(plots)


    def generate_results_parallel(self, name: str, file_path: str, modelo: str, type: int = XLSX, zip_compression: bool = False, 
                         decimals: int = 2):

        labels = self.get_labels()
        # plot_labels['simanfor.general.Summary'] = i18n.t('simanfor.general.Summary')

        plots = self.get_first_step().inventory.plots

        outfile_list = []

        for plot in plots:
            if type == XLSX:
                # recalc_list.append(self.recalculate_process(new_plot, model, 
                #         operation.get_variable('time'), result_pies_mayores, result_inventory))

                outfile_list.append(delayed(self.generate_xslt_file)(name, labels,
                    file_path + OUTPUT_FILE_BASE + str(plot.id) + '.' + OUTPUT_EXTENSION[type], 
                    plot, modelo, decimals=decimals))

        # result_inventory = (delayed(result_inventory.add_plots)(recalc_list)).compute(scheduler="distributed") # default, same as empty ""
        nbr_plots = delayed(self.print_plots)(outfile_list).compute(scheduler="distributed")
        # nbr_plots = delayed(self.print_plots)(outfile_list).compute(scheduler="single-threaded")
        print("Printed", nbr_plots, "output files.")
